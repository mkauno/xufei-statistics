# -*- coding: gbk -*-
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os

xlsx_path = r'E:/工作稿/2026.5/小学续费统计_5.15.xlsx'
xgao_path = r'E:/工作稿/2026.5/20260515小高综合2026Q2续2026Q3续班率.xls'
orig_path = r'E:/工作稿/2026.5/26年小学单科Q2数据-5.15.xlsx'

has_xgao = os.path.exists(xgao_path)

# ── 读取原始三科 ──────────────────────────────────────────────────
orig = pd.read_excel(orig_path, sheet_name=None, engine='calamine')
yw = orig['语文'].copy(); sx = orig['数学'].copy(); yy = orig['英语'].copy()
for df in [yw, sx, yy]:
    df['基数'] = pd.to_numeric(df['基数'], errors='coerce').fillna(0)
    df['是否续费3季度'] = pd.to_numeric(df['是否续费3季度'], errors='coerce').fillna(0)

# 语文排除拼音（总体合计专用）
yw_main = yw[yw['课程年级'] != '拼音'].copy()

# ── 读取综合（可选） ─────────────────────────────────────────────
xgao_e = xgao_m = None
if has_xgao:
    xgao = pd.read_excel(xgao_path, sheet_name='续班数据表', engine='openpyxl')
    xgao['5.14续班'] = pd.to_numeric(xgao['5.14续班'], errors='coerce').fillna(0)
    xgao['基数'] = pd.to_numeric(xgao['基数'], errors='coerce').fillna(0)
    xgao_e = xgao[xgao['学科'] == '素养E'].copy()
    xgao_m = xgao[xgao['学科'] == '素养M'].copy()

# ── 幼小拼音筛选 ──────────────────────────────────────────────────
yw_yx = yw[yw['课程年级'].isin(['幼小', '拼音'])].copy()
yw_yx.loc[yw_yx['课程年级'] == '幼小', '分组'] = '语文-幼小'
yw_yx.loc[yw_yx['课程年级'] == '拼音', '分组'] = '语文-拼音'
sx_yx = sx[sx['课程年级'] == '幼小'].copy(); sx_yx['分组'] = '数学-幼小'
yy_yx = yy[yy['课程年级'] == '幼小'].copy(); yy_yx['分组'] = '英语-幼小'
yx_all = pd.concat([yw_yx, sx_yx, yy_yx])
yx_b = int(yx_all['基数'].sum())
yx_x = yx_all['是否续费3季度'].sum()

# ── 科目总计 ──────────────────────────────────────────────────────
yw_b = int(yw['基数'].sum()); yw_x = yw['是否续费3季度'].sum()
sx_b = int(sx['基数'].sum()); sx_x = sx['是否续费3季度'].sum()
yy_b = int(yy['基数'].sum()); yy_x = yy['是否续费3季度'].sum()

zh_b = zh_x = 0
sx_tea_total_b = sx_tea_total_x = sx_b
yy_tea_total_b = yy_tea_total_x = yy_b

if has_xgao:
    zh_b = int(xgao_e['基数'].sum()) + int(xgao_m['基数'].sum())
    zh_x = xgao_e['5.14续班'].sum() + xgao_m['5.14续班'].sum()
    sx_tea_total_b = sx_b + int(xgao_m['基数'].sum())
    sx_tea_total_x = sx_x + xgao_m['5.14续班'].sum()
    yy_tea_total_b = yy_b + int(xgao_e['基数'].sum())
    yy_tea_total_x = yy_x + xgao_e['5.14续班'].sum()

# ── 辅助 ───────────────────────────────────────────────────────────
def is_int(v):
    try: f = float(v); return f == int(f) if f == f else False
    except: return False
def fmt_xf(v):
    if v is None: return 0
    return int(v) if is_int(v) else round(float(v), 1)

def hdr(ws, r, c, val, bg='4472C4', fg='FFFFFF', bold=True, sz=11):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = Font(bold=bold, color=fg, size=sz, name='Arial')
    cell.fill = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center')

def dc(ws, r, c, val, fmt=None, bold=False, align='center', bg=None):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = Font(bold=bold, size=10, name='Arial')
    if bg: cell.fill = PatternFill('solid', fgColor=bg)
    if fmt: cell.number_format = fmt
    cell.alignment = Alignment(horizontal=align, vertical='center')

def write_title(ws, r, title, bg='70AD47'):
    ws.merge_cells(f'A{r}:D{r}')
    c = ws.cell(row=r, column=1, value=title)
    c.font = Font(bold=True, color='FFFFFF', size=11, name='Arial')
    c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[r].height = 22
    return r + 1

def write_hdr(ws, r, hdrs, bg='375623'):
    for ci, h in enumerate(hdrs, 1):
        hdr(ws, r, ci, h, bg=bg)
    ws.row_dimensions[r].height = 20
    return r + 1

def write_rows(ws, r, df, name_col, tb, tx):
    for idx, (_, row) in enumerate(df.iterrows()):
        bg_c = 'EBF1DE' if idx % 2 == 0 else 'FFFFFF'
        lv = row['续费'] / row['基数'] if row['基数'] > 0 else 0
        dc(ws, r, 1, row[name_col], align='left', bg=bg_c)
        dc(ws, r, 2, fmt_xf(row['基数']), align='center', bg=bg_c)
        dc(ws, r, 3, fmt_xf(row['续费']), align='center', bg=bg_c)
        dc(ws, r, 4, lv, fmt='0.0%', align='center', bg=bg_c)
        ws.row_dimensions[r].height = 18
        r += 1
    tlv = tx / tb if tb > 0 else 0
    dc(ws, r, 1, '合计', bold=True, align='center', bg='C6EFCE')
    dc(ws, r, 2, fmt_xf(tb), bold=True, align='center', bg='C6EFCE')
    dc(ws, r, 3, fmt_xf(tx), bold=True, align='center', bg='C6EFCE')
    dc(ws, r, 4, tlv, fmt='0.0%', bold=True, align='center', bg='C6EFCE')
    ws.row_dimensions[r].height = 20
    return r + 2

def build_summary_start(ws, title_text):
    ws.merge_cells('A1:D1')
    hdr(ws, 1, 1, title_text, bg='1F3864', sz=14)
    ws.row_dimensions[1].height = 30
    for ci, h in enumerate(['科目/分组', '基数', '续费', '续费率'], 1):
        hdr(ws, 2, ci, h)

# ── 打开/新建xlsx ────────────────────────────────────────────────
if os.path.exists(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path)
    for sn in ['幼小拼音汇总', '总体合计汇总']:
        if sn in wb.sheetnames:
            del wb[sn]
else:
    wb = openpyxl.Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

# ═══════════════════════════════════════════════════════════════════
# Sheet 1: 幼小拼音汇总
# ═══════════════════════════════════════════════════════════════════
ws1 = wb.create_sheet('幼小拼音汇总')
build_summary_start(ws1, '幼小拼音汇总')

groups_display = [('语文-幼小', int(yw_yx[yw_yx['分组']=='语文-幼小']['基数'].sum()), yw_yx[yw_yx['分组']=='语文-幼小']['是否续费3季度'].sum()),
                  ('语文-拼音', int(yw_yx[yw_yx['分组']=='语文-拼音']['基数'].sum()), yw_yx[yw_yx['分组']=='语文-拼音']['是否续费3季度'].sum()),
                  ('数学-幼小', int(sx_yx['基数'].sum()), sx_yx['是否续费3季度'].sum()),
                  ('英语-幼小', int(yy_yx['基数'].sum()), yy_yx['是否续费3季度'].sum())]
groups_display.sort(key=lambda s: s[2]/s[1] if s[1] > 0 else 0, reverse=True)

r = 3
for subj, b, x in groups_display:
    lv = x / b if b > 0 else 0
    bg_c = 'D6DCE5' if (r - 2) % 2 == 0 else 'FFFFFF'
    dc(ws1, r, 1, subj, bold=True, align='center', bg=bg_c)
    dc(ws1, r, 2, fmt_xf(b), align='center', bg=bg_c)
    dc(ws1, r, 3, fmt_xf(x), align='center', bg=bg_c)
    dc(ws1, r, 4, lv, fmt='0.0%', align='center', bg=bg_c)
    ws1.row_dimensions[r].height = 20
    r += 1

dc(ws1, r, 1, '幼小拼音总计', bold=True, align='center', bg='FFC000')
dc(ws1, r, 2, fmt_xf(yx_b), bold=True, align='center', bg='FFC000')
dc(ws1, r, 3, fmt_xf(yx_x), bold=True, align='center', bg='FFC000')
dc(ws1, r, 4, yx_x / yx_b if yx_b > 0 else 0, fmt='0.0%', bold=True, align='center', bg='FFC000')
ws1.row_dimensions[r].height = 20
r += 3

for grp_name, grp_df in [('语文-幼小', yw_yx[yw_yx['分组']=='语文-幼小']),
                         ('语文-拼音', yw_yx[yw_yx['分组']=='语文-拼音']),
                         ('数学-幼小', sx_yx),
                         ('英语-幼小', yy_yx)]:
    grp_b = int(grp_df['基数'].sum()); grp_x = grp_df['是否续费3季度'].sum()
    tea = grp_df.groupby('任课老师', as_index=False).agg(基数=('基数', 'sum'), 续费=('是否续费3季度', 'sum'))
    tea['续费率'] = tea['续费'] / tea['基数']
    tea = tea.sort_values('续费率', ascending=False).reset_index(drop=True)
    r = write_title(ws1, r, f'{grp_name} 教师续费汇总', bg='70AD47')
    r = write_hdr(ws1, r, ['任课老师', '基数', '续费', '续费率'])
    r = write_rows(ws1, r, tea, '任课老师', grp_b, grp_x)

yx_qy = yx_all.groupby('上课校区', as_index=False).agg(基数=('基数', 'sum'), 续费=('是否续费3季度', 'sum'))
yx_qy['续费率'] = yx_qy['续费'] / yx_qy['基数']
yx_qy = yx_qy.sort_values('续费率', ascending=False).reset_index(drop=True)
r = write_title(ws1, r, '幼小拼音校区续费汇总（全合并）', bg='4472C4')
r = write_hdr(ws1, r, ['校区', '基数', '续费', '续费率'], bg='1F3864')
r = write_rows(ws1, r, yx_qy, '上课校区', yx_b, yx_x)

ws1.column_dimensions['A'].width = 24
ws1.column_dimensions['B'].width = 12
ws1.column_dimensions['C'].width = 12
ws1.column_dimensions['D'].width = 12
ws1.freeze_panes = 'A3'

# ═══════════════════════════════════════════════════════════════════
# Sheet 2: 总体合计汇总
# ═══════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet('总体合计汇总')
ws2.merge_cells('A1:D1')
hdr(ws2, 1, 1, '总体合计汇总', bg='1F3864', sz=14)
ws2.row_dimensions[1].height = 30
for ci, h in enumerate(['科目', '基数', '续费', '续费率'], 1):
    hdr(ws2, 2, ci, h)

# 总体数据
if has_xgao:
    subjs = [('英语', yy_b, yy_x), ('语文', yw_b, yw_x), ('数学', sx_b, sx_x), ('小高综合', zh_b, zh_x)]
else:
    subjs = [('英语', yy_b, yy_x), ('语文', yw_b, yw_x), ('数学', sx_b, sx_x)]
subjs.sort(key=lambda s: s[2]/s[1] if s[1] > 0 else 0, reverse=True)

for ri, (subj, b, x) in enumerate(subjs, 3):
    lv = x / b if b > 0 else 0
    bg_c = 'D6DCE5' if ri % 2 == 0 else 'FFFFFF'
    dc(ws2, ri, 1, subj, bold=True, align='center', bg=bg_c)
    dc(ws2, ri, 2, fmt_xf(b), align='center', bg=bg_c)
    dc(ws2, ri, 3, fmt_xf(x), align='center', bg=bg_c)
    dc(ws2, ri, 4, lv, fmt='0.0%', align='center', bg=bg_c)
    ws2.row_dimensions[ri].height = 20

# 总计行
align_total = yw_b + sx_b + yy_b + (zh_b if has_xgao else 0)
align_total_x = yw_x + sx_x + yy_x + (zh_x if has_xgao else 0)
total_label = '四科总计' if has_xgao else '三科总计'
total_row = 3 + len(subjs)

dc(ws2, total_row, 1, total_label, bold=True, align='center', bg='FFC000')
dc(ws2, total_row, 2, fmt_xf(align_total), bold=True, align='center', bg='FFC000')
dc(ws2, total_row, 3, fmt_xf(align_total_x), bold=True, align='center', bg='FFC000')
dc(ws2, total_row, 4, align_total_x / align_total if align_total > 0 else 0, fmt='0.0%', bold=True, align='center', bg='FFC000')
ws2.row_dimensions[total_row].height = 20

# ── 校区 ──────────────────────────────────────────────────────────
yw_qy = yw.groupby('上课校区', as_index=False).agg(基数=('基数', 'sum'), 续费=('是否续费3季度', 'sum'))
sx_qy = sx.groupby('上课校区', as_index=False).agg(基数=('基数', 'sum'), 续费=('是否续费3季度', 'sum'))
yy_qy = yy.groupby('上课校区', as_index=False).agg(基数=('基数', 'sum'), 续费=('是否续费3季度', 'sum'))
qy_parts = [yw_qy, sx_qy, yy_qy]

if has_xgao:
    zh_qy = pd.concat([
        xgao_e.groupby('上课校区', as_index=False).agg(基数=('基数', 'sum'), 续费=('5.14续班', 'sum')),
        xgao_m.groupby('上课校区', as_index=False).agg(基数=('基数', 'sum'), 续费=('5.14续班', 'sum'))
    ]).groupby('上课校区', as_index=False).agg(基数=('基数', 'sum'), 续费=('续费', 'sum'))
    qy_parts.append(zh_qy)

all_qy = pd.concat(qy_parts).groupby('上课校区', as_index=False).agg(基数=('基数', 'sum'), 续费=('续费', 'sum'))
all_qy['续费率'] = all_qy['续费'] / all_qy['基数']
all_qy = all_qy.sort_values('续费率', ascending=False).reset_index(drop=True)
all_qy_b = int(all_qy['基数'].sum())
all_qy_x = all_qy['续费'].sum()

# ── 教师数据 ──────────────────────────────────────────────────────
yw_tea = yw_main.groupby('任课老师', as_index=False).agg(基数=('基数', 'sum'), 续费=('是否续费3季度', 'sum'))
yw_tea['续费率'] = yw_tea['续费'] / yw_tea['基数']
yw_tea = yw_tea.sort_values('续费率', ascending=False).reset_index(drop=True)

sx_tea = sx.groupby('任课老师', as_index=False).agg(基数=('基数', 'sum'), 续费=('是否续费3季度', 'sum'))
if has_xgao:
    sx_tea = pd.concat([sx_tea, xgao_m.groupby('任课老师', as_index=False).agg(基数=('基数', 'sum'), 续费=('5.14续班', 'sum'))])
sx_tea = sx_tea.groupby('任课老师', as_index=False).agg(基数=('基数', 'sum'), 续费=('续费', 'sum'))
sx_tea['续费率'] = sx_tea['续费'] / sx_tea['基数']
sx_tea = sx_tea.sort_values('续费率', ascending=False).reset_index(drop=True)

yy_tea = yy.groupby('任课老师', as_index=False).agg(基数=('基数', 'sum'), 续费=('是否续费3季度', 'sum'))
if has_xgao:
    yy_tea = pd.concat([yy_tea, xgao_e.groupby('任课老师', as_index=False).agg(基数=('基数', 'sum'), 续费=('5.14续班', 'sum'))])
yy_tea = yy_tea.groupby('任课老师', as_index=False).agg(基数=('基数', 'sum'), 续费=('续费', 'sum'))
yy_tea['续费率'] = yy_tea['续费'] / yy_tea['基数']
yy_tea = yy_tea.sort_values('续费率', ascending=False).reset_index(drop=True)

# ── 打印校验 ──────────────────────────────────────────────────────
print(f'=== 数据校验 ===')
print(f'幼小拼音: {yx_b}/{yx_x}/{yx_x/yx_b:.1%}')
print(f'')
print(f'语文: {yw_b}/{yw_x}/{yw_x/yw_b:.1%}')
print(f'数学: {sx_b}/{sx_x}/{sx_x/sx_b:.1%}')
print(f'英语: {yy_b}/{yy_x}/{yy_x/yy_b:.1%}')
if has_xgao:
    print(f'小高综合: {zh_b}/{zh_x}/{zh_x/zh_b:.1%}')
    print(f'四科总计: {align_total}/{align_total_x}/{align_total_x/align_total:.1%}')
else:
    print(f'(无综合文件)')
    print(f'三科总计: {align_total}/{align_total_x}/{align_total_x/align_total:.1%}')
print(f'')
print(f'校区: {len(all_qy)}个')
print(f'语文教师: {len(yw_tea)}人')
sx_label = '数学教师(含素养M)' if has_xgao else '数学教师'
yy_label = '英语教师(含素养E)' if has_xgao else '英语教师'
print(f'{sx_label}: {len(sx_tea)}人 | {yy_label}: {len(yy_tea)}人')

# ── 写总体合计汇总 ───────────────────────────────────────────────
r = total_row + 2
r = write_title(ws2, r, '校区续费汇总（全学科）', bg='4472C4')
r = write_hdr(ws2, r, ['校区', '基数', '续费', '续费率'], bg='1F3864')
r = write_rows(ws2, r, all_qy, '上课校区', all_qy_b, all_qy_x)

r = write_title(ws2, r, '语文教师续费汇总', bg='70AD47')
r = write_hdr(ws2, r, ['任课老师', '基数', '续费', '续费率'])
r = write_rows(ws2, r, yw_tea, '任课老师', yw_b, yw_x)

sx_title = '数学教师续费汇总（含小高综合素养M）' if has_xgao else '数学教师续费汇总'
r = write_title(ws2, r, sx_title, bg='70AD47')
r = write_hdr(ws2, r, ['任课老师', '基数', '续费', '续费率'])
r = write_rows(ws2, r, sx_tea, '任课老师', sx_tea_total_b, sx_tea_total_x)

yy_title = '英语教师续费汇总（含小高综合素养E）' if has_xgao else '英语教师续费汇总'
r = write_title(ws2, r, yy_title, bg='70AD47')
r = write_hdr(ws2, r, ['任课老师', '基数', '续费', '续费率'])
r = write_rows(ws2, r, yy_tea, '任课老师', yy_tea_total_b, yy_tea_total_x)

ws2.column_dimensions['A'].width = 30
ws2.column_dimensions['B'].width = 12
ws2.column_dimensions['C'].width = 12
ws2.column_dimensions['D'].width = 12
ws2.freeze_panes = 'A3'

wb.save(xlsx_path)
print(f'\n已保存: {xlsx_path}')
print(f'Sheet1: 幼小拼音汇总 | Sheet2: 总体合计汇总')
print(f'综合文件: {"有" if has_xgao else "无"}')
